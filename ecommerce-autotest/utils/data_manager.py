#!/usr/bin/env python3
"""
测试数据管理模块
提供多种数据源（YAML、JSON、Excel）的测试数据加载和管理功能
支持随机数据生成和测试数据验证
"""

import os
import json
import yaml
import random
import string
from typing import Dict, List, Any, Optional, Union, Tuple
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl import Workbook, load_workbook

from utils.logger import get_logger
from utils.common import file_utils, string_utils, validation_utils


class TestDataManager:
    """
    测试数据管理器
    支持多种数据格式的加载和管理
    """

    def __init__(self, data_source: str = 'yaml', data_dir: Optional[str] = None):
        """
        初始化测试数据管理器

        Args:
            data_source: 数据源类型，支持 'yaml', 'json', 'excel'
            data_dir: 测试数据目录路径
        """
        self.data_source = data_source.lower()
        self.data_dir = data_dir or self._get_default_data_dir()
        self.logger = get_logger(self.__class__.__name__)

        # 支持的数据源
        self.supported_sources = ['yaml', 'json', 'excel']

        if self.data_source not in self.supported_sources:
            raise ValueError(f"不支持的数据源: {self.data_source}，支持: {self.supported_sources}")

        # 确保数据目录存在
        file_utils.ensure_directory(self.data_dir)

        self.logger.info(f"测试数据管理器初始化完成，数据源: {self.data_source}，目录: {self.data_dir}")

    def _get_default_data_dir(self) -> str:
        """获取默认数据目录"""
        project_root = Path(__file__).parent.parent
        return str(project_root / 'config' / 'testdata')

    def load_test_data(self, module_name: str, data_file: Optional[str] = None) -> Dict[str, Any]:
        """
        加载测试数据

        Args:
            module_name: 模块名称，如 'login', 'product'
            data_file: 数据文件路径（可选），如果提供则使用指定文件

        Returns:
            Dict[str, Any]: 测试数据字典
        """
        try:
            if data_file:
                # 使用指定的数据文件
                file_path = data_file
                if not os.path.exists(file_path):
                    raise FileNotFoundError(f"数据文件不存在: {file_path}")
            else:
                # 根据数据源类型构建文件路径
                file_path = self._get_data_file_path(module_name)

            self.logger.debug(f"加载测试数据: {file_path}")

            # 根据文件扩展名选择加载方法
            if file_path.endswith('.yaml') or file_path.endswith('.yml'):
                data = self._load_yaml_data(file_path)
            elif file_path.endswith('.json'):
                data = self._load_json_data(file_path)
            elif file_path.endswith('.xlsx') or file_path.endswith('.xls'):
                data = self._load_excel_data(file_path, module_name)
            else:
                # 根据初始化时的数据源类型加载
                if self.data_source == 'yaml':
                    data = self._load_yaml_data(file_path)
                elif self.data_source == 'json':
                    data = self._load_json_data(file_path)
                elif self.data_source == 'excel':
                    data = self._load_excel_data(file_path, module_name)
                else:
                    raise ValueError(f"不支持的数据文件格式: {file_path}")

            self.logger.info(f"测试数据加载成功: {module_name}，数据量: {len(data) if isinstance(data, dict) else 'N/A'}")
            return data

        except Exception as e:
            self.logger.error(f"加载测试数据失败: {module_name}，错误: {e}")
            raise

    def _get_data_file_path(self, module_name: str) -> str:
        """
        获取数据文件路径

        Args:
            module_name: 模块名称

        Returns:
            str: 数据文件路径
        """
        if self.data_source == 'yaml':
            filename = f"{module_name}.yaml"
        elif self.data_source == 'json':
            filename = f"{module_name}.json"
        elif self.data_source == 'excel':
            filename = "test_data.xlsx"
        else:
            filename = f"{module_name}.yaml"

        file_path = os.path.join(self.data_dir, filename)

        # 如果文件不存在，尝试其他扩展名
        if not os.path.exists(file_path):
            alternative_paths = [
                os.path.join(self.data_dir, f"{module_name}.yml"),
                os.path.join(self.data_dir, f"{module_name}.json"),
                os.path.join(self.data_dir, "test_data.xlsx"),
            ]

            for alt_path in alternative_paths:
                if os.path.exists(alt_path):
                    self.logger.debug(f"使用备用文件: {alt_path}")
                    return alt_path

            # 如果所有文件都不存在，创建默认数据文件
            self.logger.warning(f"数据文件不存在，创建默认文件: {file_path}")
            self._create_default_data_file(file_path, module_name)

        return file_path

    def _load_yaml_data(self, file_path: str) -> Dict[str, Any]:
        """加载YAML数据"""
        with open(file_path, 'r', encoding='utf-8') as f:
            return yaml.safe_load(f) or {}

    def _load_json_data(self, file_path: str) -> Dict[str, Any]:
        """加载JSON数据"""
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)

    def _load_excel_data(self, file_path: str, sheet_name: str) -> Dict[str, Any]:
        """加载Excel数据"""
        try:
            workbook = load_workbook(file_path, data_only=True)

            # 尝试使用模块名作为工作表名，如果不存在则使用第一个工作表
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
            else:
                sheet = workbook[workbook.sheetnames[0]]
                self.logger.warning(f"工作表 '{sheet_name}' 不存在，使用 '{sheet.title}'")

            # 读取表头
            headers = []
            for cell in sheet[1]:
                headers.append(cell.value)

            # 读取数据行
            data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for i, header in enumerate(headers):
                    if i < len(row):
                        row_dict[header] = row[i]
                data.append(row_dict)

            return {"data": data, "headers": headers}

        except Exception as e:
            self.logger.error(f"加载Excel数据失败: {e}")
            return {"data": [], "headers": []}

    def _create_default_data_file(self, file_path: str, module_name: str) -> None:
        """创建默认数据文件"""
        default_data = self._get_default_test_data(module_name)

        # 根据文件扩展名保存数据
        if file_path.endswith('.yaml') or file_path.endswith('.yml'):
            with open(file_path, 'w', encoding='utf-8') as f:
                yaml.dump(default_data, f, default_flow_style=False, allow_unicode=True)
        elif file_path.endswith('.json'):
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(default_data, f, ensure_ascii=False, indent=2)
        elif file_path.endswith('.xlsx'):
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = module_name

            # 写入表头和数据
            if default_data and isinstance(default_data, dict):
                # TODO: 根据数据结构写入Excel
                pass

            workbook.save(file_path)

        self.logger.info(f"默认数据文件已创建: {file_path}")

    def _get_default_test_data(self, module_name: str) -> Dict[str, Any]:
        """获取默认测试数据"""
        default_data = {
            'login': {
                'valid_users': [
                    {
                        'username': 'admin',
                        'password': 'password123',
                        'role': '管理员'
                    },
                    {
                        'username': 'operator',
                        'password': 'operator123',
                        'role': '操作员'
                    }
                ],
                'invalid_users': [
                    {
                        'username': '',
                        'password': 'password123',
                        'expected_error': '用户名不能为空'
                    },
                    {
                        'username': 'admin',
                        'password': '',
                        'expected_error': '密码不能为空'
                    },
                    {
                        'username': 'wrong',
                        'password': 'wrong',
                        'expected_error': '账号或密码错误'
                    }
                ]
            },
            'product': {
                'categories': ['电子产品', '服装', '食品', '家居', '图书'],
                'test_products': [
                    {
                        'name': '智能手机',
                        'category': '电子产品',
                        'price': 2999.99,
                        'stock': 100,
                        'description': '高性能智能手机'
                    },
                    {
                        'name': '笔记本电脑',
                        'category': '电子产品',
                        'price': 5999.99,
                        'stock': 50,
                        'description': '轻薄笔记本电脑'
                    }
                ]
            },
            'user': {
                'test_users': [
                    {
                        'username': 'test_user_1',
                        'email': 'user1@example.com',
                        'phone': '13800138001',
                        'role': '普通用户'
                    },
                    {
                        'username': 'test_user_2',
                        'email': 'user2@example.com',
                        'phone': '13800138002',
                        'role': 'VIP用户'
                    }
                ]
            }
        }

        return default_data.get(module_name, {})

    def save_test_data(self, module_name: str, data: Dict[str, Any], data_file: Optional[str] = None) -> bool:
        """
        保存测试数据

        Args:
            module_name: 模块名称
            data: 测试数据
            data_file: 数据文件路径（可选）

        Returns:
            bool: 是否保存成功
        """
        try:
            if data_file:
                file_path = data_file
            else:
                file_path = self._get_data_file_path(module_name)

            # 根据文件扩展名选择保存方法
            if file_path.endswith('.yaml') or file_path.endswith('.yml'):
                with open(file_path, 'w', encoding='utf-8') as f:
                    yaml.dump(data, f, default_flow_style=False, allow_unicode=True)
            elif file_path.endswith('.json'):
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
            elif file_path.endswith('.xlsx') or file_path.endswith('.xls'):
                self.logger.warning("Excel数据保存功能尚未实现")
                return False

            self.logger.info(f"测试数据保存成功: {file_path}")
            return True

        except Exception as e:
            self.logger.error(f"保存测试数据失败: {e}")
            return False

    def generate_random_data(self, data_template: Dict[str, Any], count: int = 1) -> List[Dict[str, Any]]:
        """
        生成随机测试数据

        Args:
            data_template: 数据模板，定义每个字段的生成规则
            count: 生成的数据条数

        Returns:
            List[Dict[str, Any]]: 随机生成的测试数据列表
        """
        random_data = []

        for i in range(count):
            item = {}
            for field, template in data_template.items():
                item[field] = self._generate_random_field(template, i + 1)
            random_data.append(item)

        self.logger.debug(f"生成 {count} 条随机测试数据")
        return random_data

    def _generate_random_field(self, template: Any, index: int) -> Any:
        """生成随机字段值"""
        if isinstance(template, dict):
            # 模板包含生成规则
            field_type = template.get('type', 'string')
            options = template.get('options')
            min_val = template.get('min', 1)
            max_val = template.get('max', 100)
            prefix = template.get('prefix', '')
            suffix = template.get('suffix', '')

            if field_type == 'string':
                if options:
                    # 从选项列表中随机选择
                    return random.choice(options)
                else:
                    # 生成随机字符串
                    length = random.randint(min_val, max(max_val, min_val + 10))
                    random_str = string_utils.generate_random_string(length)
                    return f"{prefix}{random_str}{suffix}"

            elif field_type == 'number':
                # 生成随机数字
                if isinstance(min_val, int) and isinstance(max_val, int):
                    return random.randint(min_val, max_val)
                else:
                    return round(random.uniform(min_val, max_val), 2)

            elif field_type == 'email':
                # 生成随机邮箱
                username = string_utils.generate_random_string(8)
                domain = template.get('domain', 'example.com')
                return f"{username}@{domain}"

            elif field_type == 'phone':
                # 生成随机手机号
                return string_utils.generate_random_phone()

            elif field_type == 'date':
                # 生成随机日期
                start_date = template.get('start_date', '2020-01-01')
                end_date = template.get('end_date', '2024-12-31')

                try:
                    start = datetime.strptime(start_date, '%Y-%m-%d')
                    end = datetime.strptime(end_date, '%Y-%m-%d')
                    delta = end - start
                    random_days = random.randint(0, delta.days)
                    random_date = start + timedelta(days=random_days)
                    return random_date.strftime('%Y-%m-%d')
                except ValueError:
                    return datetime.now().strftime('%Y-%m-%d')

            elif field_type == 'boolean':
                # 生成随机布尔值
                return random.choice([True, False])

            else:
                # 未知类型，返回模板本身
                return template

        elif isinstance(template, list):
            # 从列表中随机选择
            return random.choice(template)

        else:
            # 直接使用模板值，但如果是字符串包含 {index} 则替换
            if isinstance(template, str):
                return template.replace('{index}', str(index))
            return template

    def get_test_case_data(self, module_name: str, test_case: str) -> List[Dict[str, Any]]:
        """
        获取测试用例数据

        Args:
            module_name: 模块名称
            test_case: 测试用例名称

        Returns:
            List[Dict[str, Any]]: 测试用例数据列表
        """
        try:
            data = self.load_test_data(module_name)

            # 支持多种数据结构
            if test_case in data:
                test_data = data[test_case]
            elif 'test_cases' in data and test_case in data['test_cases']:
                test_data = data['test_cases'][test_case]
            elif 'data' in data:
                test_data = data['data']
            else:
                self.logger.warning(f"未找到测试用例数据: {module_name}.{test_case}")
                return []

            # 确保返回列表
            if isinstance(test_data, dict):
                return [test_data]
            elif isinstance(test_data, list):
                return test_data
            else:
                return [{'value': test_data}]

        except Exception as e:
            self.logger.error(f"获取测试用例数据失败: {module_name}.{test_case}，错误: {e}")
            return []

    def validate_test_data(self, module_name: str, data: Dict[str, Any]) -> Tuple[bool, List[str]]:
        """
        验证测试数据

        Args:
            module_name: 模块名称
            data: 测试数据

        Returns:
            Tuple[bool, List[str]]: (是否有效, 错误信息列表)
        """
        errors = []

        # 模块特定的验证规则
        validation_rules = self._get_validation_rules(module_name)

        for field, rules in validation_rules.items():
            if field in data:
                value = data[field]
                for rule in rules:
                    error = self._validate_field(field, value, rule)
                    if error:
                        errors.append(error)

        return len(errors) == 0, errors

    def _get_validation_rules(self, module_name: str) -> Dict[str, List[str]]:
        """获取验证规则"""
        rules = {
            'login': {
                'username': ['required', 'string'],
                'password': ['required', 'string'],
                'expected_error': ['string']
            },
            'product': {
                'name': ['required', 'string'],
                'price': ['required', 'number'],
                'stock': ['required', 'number']
            },
            'user': {
                'username': ['required', 'string'],
                'email': ['required', 'email'],
                'phone': ['phone']
            }
        }

        return rules.get(module_name, {})

    def _validate_field(self, field: str, value: Any, rule: str) -> Optional[str]:
        """验证字段"""
        if rule == 'required':
            if value is None or (isinstance(value, str) and value.strip() == ''):
                return f"字段 '{field}' 是必填项"
        elif rule == 'string':
            if not isinstance(value, str):
                return f"字段 '{field}' 必须是字符串类型"
        elif rule == 'number':
            if not isinstance(value, (int, float)):
                return f"字段 '{field}' 必须是数字类型"
        elif rule == 'email':
            if not string_utils.is_valid_email(value):
                return f"字段 '{field}' 必须是有效的邮箱地址"
        elif rule == 'phone':
            if not string_utils.is_valid_phone(value):
                return f"字段 '{field}' 必须是有效的手机号"

        return None

    def merge_test_data(self, base_data: Dict[str, Any], override_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        合并测试数据

        Args:
            base_data: 基础数据
            override_data: 覆盖数据

        Returns:
            Dict[str, Any]: 合并后的数据
        """
        from utils.common import data_utils
        return data_utils.deep_merge(base_data, override_data)


# 全局测试数据管理器实例
_test_data_manager = None


def get_test_data_manager(data_source: str = 'yaml', data_dir: Optional[str] = None) -> TestDataManager:
    """
    获取测试数据管理器实例（单例模式）

    Args:
        data_source: 数据源类型
        data_dir: 数据目录

    Returns:
        TestDataManager: 测试数据管理器实例
    """
    global _test_data_manager

    if _test_data_manager is None:
        _test_data_manager = TestDataManager(data_source, data_dir)

    return _test_data_manager


def load_test_data(module_name: str, data_file: Optional[str] = None) -> Dict[str, Any]:
    """
    加载测试数据的快捷函数

    Args:
        module_name: 模块名称
        data_file: 数据文件路径

    Returns:
        Dict[str, Any]: 测试数据
    """
    manager = get_test_data_manager()
    return manager.load_test_data(module_name, data_file)


def generate_random_data(data_template: Dict[str, Any], count: int = 1) -> List[Dict[str, Any]]:
    """
    生成随机测试数据的快捷函数

    Args:
        data_template: 数据模板
        count: 生成数量

    Returns:
        List[Dict[str, Any]]: 随机测试数据
    """
    manager = get_test_data_manager()
    return manager.generate_random_data(data_template, count)


if __name__ == "__main__":
    # 测试测试数据管理器
    print("测试测试数据管理器...")

    # 创建测试数据管理器
    manager = TestDataManager(data_source='yaml')

    # 测试加载默认数据
    try:
        login_data = manager.load_test_data('login')
        print(f"登录测试数据: {login_data.keys() if isinstance(login_data, dict) else 'N/A'}")

        # 测试随机数据生成
        template = {
            'username': {'type': 'string', 'min': 5, 'max': 10},
            'email': {'type': 'email', 'domain': 'test.com'},
            'age': {'type': 'number', 'min': 18, 'max': 60},
            'is_active': {'type': 'boolean'}
        }

        random_data = manager.generate_random_data(template, 3)
        print(f"生成的随机数据: {random_data}")

        # 测试数据验证
        test_user = {
            'username': 'testuser',
            'email': 'test@example.com',
            'phone': '13800138000'
        }

        valid, errors = manager.validate_test_data('user', test_user)
        print(f"数据验证结果: 有效={valid}, 错误={errors}")

        print("\n测试数据管理器测试完成")

    except Exception as e:
        print(f"测试失败: {e}")
        import traceback
        traceback.print_exc()